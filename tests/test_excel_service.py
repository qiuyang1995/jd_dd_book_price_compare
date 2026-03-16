import tempfile
import unittest
from pathlib import Path

import openpyxl

from price_app.excel_service import ExcelPriceWorkbook, WorkbookStructureError


class ExcelPriceWorkbookTests(unittest.TestCase):
    def test_workbook_creates_output_columns_and_writes_results(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "books.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet["A1"] = "ISBN"
            sheet["A2"] = "9787308262453"
            workbook.save(workbook_path)

            price_workbook = ExcelPriceWorkbook(workbook_path)
            rows = list(price_workbook.iter_isbn_rows())
            self.assertEqual(len(rows), 1)
            price_workbook.write_result(rows[0].row_index, "39.90", "29.80", "满100减20")
            price_workbook.save()

            reloaded = openpyxl.load_workbook(workbook_path).active
            self.assertEqual(reloaded["B1"].value, "京东价格")
            self.assertEqual(reloaded["C1"].value, "当当价格")
            self.assertEqual(reloaded["D1"].value, "当当优惠")
            self.assertEqual(reloaded["B2"].value, "39.90")
            self.assertEqual(reloaded["C2"].value, "29.80")
            self.assertEqual(reloaded["D2"].value, "满100减20")

    def test_workbook_requires_isbn_column(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "books.xlsx"
            workbook = openpyxl.Workbook()
            workbook.active["A1"] = "书名"
            workbook.save(workbook_path)

            with self.assertRaises(WorkbookStructureError):
                ExcelPriceWorkbook(workbook_path)


if __name__ == "__main__":
    unittest.main()

