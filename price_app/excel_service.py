from __future__ import annotations

from pathlib import Path
from typing import Iterable

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from price_app.models import WorkbookColumns, WorkbookRow


class WorkbookStructureError(ValueError):
    """Excel 模板结构不满足处理要求。"""


class ExcelPriceWorkbook:
    """封装 Excel 读写，避免业务流程直接操作 openpyxl 细节。"""

    def __init__(self, file_path: str | Path):
        self.file_path = Path(file_path)
        self.workbook = openpyxl.load_workbook(self.file_path)
        self.sheet: Worksheet = self.workbook.active
        self.columns = WorkbookColumns(
            isbn=self._find_isbn_column(),
            jd_price=self._find_or_create_output_column("京东价格"),
            dd_price=self._find_or_create_output_column("当当价格"),
            dd_discount=self._find_or_create_output_column("当当优惠"),
        )

    @property
    def total_rows(self) -> int:
        return sum(1 for _ in self.iter_isbn_rows())

    def iter_isbn_rows(self) -> Iterable[WorkbookRow]:
        for row_index in range(2, self.sheet.max_row + 1):
            raw_value = self.sheet.cell(row=row_index, column=self.columns.isbn).value
            isbn = "" if raw_value is None else str(raw_value).strip()
            if isbn:
                yield WorkbookRow(row_index=row_index, isbn=isbn)

    def write_result(self, row_index: int, jd_price: str, dd_price: str, dd_discount: str) -> None:
        self.sheet.cell(row=row_index, column=self.columns.jd_price).value = jd_price
        self.sheet.cell(row=row_index, column=self.columns.dd_price).value = dd_price
        self.sheet.cell(row=row_index, column=self.columns.dd_discount).value = dd_discount

    def save(self) -> None:
        self.workbook.save(self.file_path)

    def _find_isbn_column(self) -> int:
        for column in range(1, self.sheet.max_column + 1):
            value = self.sheet.cell(row=1, column=column).value
            normalized = "" if value is None else str(value).strip().lower()
            if normalized in {"isbn", "isbn号"}:
                return column
        raise WorkbookStructureError("未找到名为 ISBN 或 ISBN号 的列")

    def _find_or_create_output_column(self, header: str) -> int:
        for column in range(1, self.sheet.max_column + 1):
            value = self.sheet.cell(row=1, column=column).value
            if value is not None and str(value).strip() == header:
                return column

        column = self.sheet.max_column + 1
        self.sheet.cell(row=1, column=column).value = header
        return column

