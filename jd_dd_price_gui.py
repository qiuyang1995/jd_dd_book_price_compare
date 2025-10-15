import tkinter as tk
import random
from tkinter import filedialog, messagebox, ttk, simpledialog
import requests
import openpyxl
import time
import threading
import os
import re
import pickle
from bs4 import BeautifulSoup
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException

COOKIE_FILE_JD = "jd_cookie.txt"
COOKIE_FILE_DD = "dd_cookie.txt"
JD_COOKIES_PKL = "jd_cookies.pkl"


class JDPriceFetcherApp:
    def __init__(self, root):
        self.root = root
        self.root.title("📚 图书价格抓取工具（京东 + 当当）")
        self.root.geometry("650x430")
        self.root.resizable(False, False)

        self.file_path = ""
        self.running = False
        self.driver = None
        self.access_restricted = False  # 添加访问受限标志
        self.base_sleep_time = 5  # 基础睡眠时间
        self.jd_search_page_loaded = False  # 标记京东搜索页面是否已加载

        # 先创建 UI，再加载 cookie
        self.create_ui()

    # ---------------- UI 部分 ----------------
    def create_ui(self):
        tk.Label(self.root, text="📘 图书价格获取工具（京东 + 当当）", font=("微软雅黑", 16, "bold")).pack(pady=10)

        tk.Button(self.root, text="选择 Excel 文件", command=self.select_file, width=20, bg="#2196F3", fg="white").pack(pady=5)
        
        # 添加测试按钮
        tk.Button(self.root, text="测试京东访问", command=self.test_jd_access, width=20, bg="#9C27B0", fg="white").pack(pady=3)
        tk.Button(self.root, text="测试当当访问", command=self.test_dd_access, width=20, bg="#4CAF50", fg="white").pack(pady=3)
        
        # 添加登录检测测试按钮
        # tk.Button(self.root, text="测试登录检测", command=self.test_login_detection, width=20, bg="#FF9800", fg="white").pack(pady=3)
        
        # 添加重置访问间隔按钮
        # tk.Button(self.root, text="重置访问间隔", command=self.reset_access_interval, width=20, bg="#FF5722", fg="white").pack(pady=3)
        
        self.start_btn = tk.Button(self.root, text="开始执行", command=self.start, width=20, bg="#FF9800", fg="white")
        self.start_btn.pack(pady=5)

        self.progress = ttk.Progressbar(self.root, length=400, mode='determinate')
        self.progress.pack(pady=10)
        
        # 创建带滚动条的日志框
        log_frame = tk.Frame(self.root)
        log_frame.pack(pady=10)
        
        # 创建文本框和滚动条
        self.log_box = tk.Text(log_frame, height=10, width=75, wrap='word', state='disabled', bg="#f4f4f4")
        scrollbar = tk.Scrollbar(log_frame, orient="vertical", command=self.log_box.yview)
        self.log_box.configure(yscrollcommand=scrollbar.set)
        
        # 布局
        self.log_box.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    # ---------------- 日志输出 ----------------
    def log(self, msg):
        # 安全地向日志窗写入（UI 线程）
        try:
            self.log_box.config(state='normal')
            self.log_box.insert(tk.END, msg + "\n")
            self.log_box.see(tk.END)
            self.log_box.config(state='disabled')
        except Exception:
            # 如果日志区域不可用，退回到控制台输出，防止程序崩溃
            print(msg)

    # ---------------- Excel 文件选择 ----------------
    def select_file(self):
        self.file_path = filedialog.askopenfilename(filetypes=[("Excel 文件", "*.xlsx")])
        if self.file_path:
            self.log(f"📂 已选择文件：{self.file_path}")

    # ---------------- 主流程 ----------------
    def start(self):
        # 检查Cookie设置不再必须，因为我们使用Selenium登录
        if not self.file_path:
            messagebox.showwarning("提示", "请先选择 Excel 文件！")
            return

        self.start_btn.config(state='disabled')
        self.running = True
        threading.Thread(target=self.process_excel).start()

    def process_excel(self):
        # 记录开始时间
        start_time = time.time()
        
        # 重置访问受限状态
        self.access_restricted = False
        self.base_sleep_time = 5
        self.jd_search_page_loaded = False  # 重置搜索页面状态
        self.request_count = 0  # 重置请求计数器
        
        try:
            # 初始化浏览器
            self.log("🚀 正在启动浏览器...")
            if not self.init_browser():
                self.log("❌ 浏览器启动失败")
                return
            
            # 登录京东
            self.log("🔑 正在登录京东...")
            if not self.login_or_load_jd_cookie():
                self.log("❌ 京东登录失败")
                return
            
            wb = openpyxl.load_workbook(self.file_path)
            sheet = wb.active

            # 找出 ISBN 列
            isbn_col = None
            if sheet is not None:  # 添加类型检查
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=1, column=col).value
                    if cell_value is not None and str(cell_value).strip().lower() in ("isbn", "isbn号"):
                        isbn_col = col
                        break

            if not isbn_col:
                messagebox.showerror("错误", "未找到名为 ISBN 的列")
                self.start_btn.config(state='normal')
                return

            # 新增列：京东价格、当当价格
            if sheet is not None:  # 添加类型检查
                jd_price_col = sheet.max_column + 1
                dd_price_col = jd_price_col + 1
                dd_discount_col = dd_price_col + 1
                # 添加类型忽略注释以消除警告
                sheet.cell(row=1, column=jd_price_col).value = "京东价格"  # type: ignore
                sheet.cell(row=1, column=dd_price_col).value = "当当价格"  # type: ignore
                sheet.cell(row=1, column=dd_discount_col).value = "当当优惠"  # type: ignore

                total = sheet.max_row - 1
                self.progress["maximum"] = total

                for i in range(2, sheet.max_row + 1):
                    try:
                        cell_value = sheet.cell(row=i, column=isbn_col).value
                        if cell_value is not None:
                            isbn = str(cell_value).strip()
                        else:
                            isbn = ""
                            
                        if not isbn:
                            continue

                        jd_price = self.fetch_price_jd(isbn)
                        
                        # 检查是否登录失败
                        if jd_price == "登录失败":
                            self.log(f"⚠️ 检测到登录失败，暂停处理并提示用户")
                            messagebox.showwarning("登录失败", "检测到京东登录失败，程序将暂停执行。\n请重新登录后再试。")
                            # 保存已处理的数据
                            wb.save(self.file_path)
                            self.log("💾 已保存当前进度")
                            return  # 退出处理流程
                        
                        # 检查是否访问受限
                        if jd_price == "访问受限":
                            self.handle_access_restriction()
                            self.log(f"⚠️ 检测到访问受限，暂停处理并提示用户")
                            messagebox.showwarning("访问受限", f"检测到京东访问受限，程序将增加请求间隔。\n当前请求间隔已调整为 {self.base_sleep_time} 秒。\n建议稍后再试或检查网络环境。")
                            # 保存已处理的数据
                            wb.save(self.file_path)
                            self.log("💾 已保存当前进度")
                            # 继续处理下一个，而不是退出
                            continue
                        
                        # 根据获取结果调整下次请求间隔
                        if jd_price in ["超时", None, "访问受限", "登录失败"]:
                            # 获取失败，增加下次请求间隔
                            self.base_sleep_time = min(self.base_sleep_time + 5, 60)  # 最多增加到60秒
                            self.log(f"⚠️ 获取失败，增加下次请求间隔至 {self.base_sleep_time} 秒")
                        elif self.base_sleep_time > 20:
                            # 获取成功且当前间隔较高，逐渐减少间隔
                            self.base_sleep_time = max(self.base_sleep_time - 2, 10)  # 最少20秒
                            self.log(f"✅ 获取成功，下次请求间隔调整为 {self.base_sleep_time} 秒")
                        
                        dd = self.fetch_price_dd(isbn)
                        
                        # 添加类型检查
                        dd_price = ""
                        dd_discount = ""
                        if dd is not None and isinstance(dd, dict):
                            dd_price = dd.get('price', '')
                            dd_discount = dd.get('discount', '')

                        # 添加类型忽略注释以消除警告
                        sheet.cell(row=i, column=jd_price_col).value = jd_price  # type: ignore
                        sheet.cell(row=i, column=dd_price_col).value = dd_price  # type: ignore
                        sheet.cell(row=i, column=dd_discount_col).value = dd_discount  # type: ignore

                        # 根据请求次数动态调整间隔时间
                        sleep_time = self.calculate_sleep_time()
                        sleep_time = int(sleep_time)
                        self.progress["value"] = i - 1
                        self.log(
                            f"{i - 1}/{total} ✅ {isbn} → 京东 ¥{jd_price or '未获取'} ｜ 当当 ¥{dd_price or '未获取'} {dd_discount} 下次请求：{sleep_time}秒后")

                        time.sleep(sleep_time)

                        # 在主线程安全更新UI
                        self.root.update_idletasks()
                    except Exception:
                        pass

                wb.save(self.file_path)
                
                # 计算总耗时
                end_time = time.time()
                elapsed_time = end_time - start_time
                hours, rem = divmod(elapsed_time, 3600)
                minutes, seconds = divmod(rem, 60)
                
                self.log("🎉 全部完成，结果已写入 Excel！")
                self.log(f"⏱️ 任务总耗时: {int(hours):02d}时{int(minutes):02d}分{int(seconds):02d}秒")
                messagebox.showinfo("完成", f"价格抓取完成！\n任务总耗时: {int(hours):02d}时{int(minutes):02d}分{int(seconds):02d}秒")

        except Exception as e:
            self.log(f"❌ 错误：{str(e)}")
        finally:
            # 确保关闭浏览器
            self.close_browser()
            self.start_btn.config(state='normal')

    # ---------------- 浏览器初始化和京东登录 ----------------
    def init_browser(self):
        """简化版本的浏览器初始化 - 只使用 undetected_chromedriver"""
        self.log("🚀 正在初始化浏览器...")
        
        # 只使用 undetected_chromedriver 方案
        try:
            import undetected_chromedriver as uc
            # self.log("🔄 使用 undetected_chromedriver...")
            
            # 创建选项
            options = uc.ChromeOptions()
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument("--disable-gpu")
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_argument("--disable-extensions")
            options.add_argument("--no-first-run")
            options.add_argument("--disable-default-apps")
            options.add_argument("--remote-debugging-port=9222")
            # 添加更多反检测选项
            options.add_argument("--disable-web-security")
            options.add_argument("--disable-features=VizDisplayCompositor")
            options.add_argument("--disable-plugins")
            options.add_argument("--disable-plugins-discovery")
            options.add_argument("--disable-preconnect")
            options.add_argument("--disable-ipc-flooding-protection")
            
            # 设置用户代理
            options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36")

            # 初始化驱动
            self.driver = uc.Chrome(options=options, driver_executable_path="D:/chromedriver/chromedriver.exe", version_main=None)
            self.driver.maximize_window()
            
            # 执行反检测脚本
            self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            self.driver.execute_script("window.chrome = {runtime: {}};")
            self.driver.execute_script("navigator.permissions.query = (parameters) => {return new Promise((resolve) => resolve({state: 'granted'}))};")
            
            # 测试访问
            self.driver.get("data:text/html,<html><body><h1>Undetected Test OK</h1></body></html>")
            
            self.log("✅ undetected_chromedriver 启动成功")
            return True

        except Exception as e:
            self.log(f"❌ undetected_chromedriver 启动失败: {e}")
            
            # 详细错误信息
            import traceback
            # self.log(f"详细错误：{traceback.format_exc()}")
            
            # 提供解决方案
            self.log("💡 可能的解决方案:")
            self.log("   1. 检查网络连接是否正常")
            self.log("   2. 关闭防火墙或杀毒软件")
            self.log("   3. 运行: pip install undetected-chromedriver")
            self.log("   4. 手动下载 ChromeDriver 并加入 PATH")
            
            return False

    def login_or_load_jd_cookie(self):
        """登录京东或加载已保存的Cookie"""
        if not self.driver:
            return False
            
        try:
            self.driver.get("https://www.jd.com/")
            
            # 重置访问受限状态
            self.access_restricted = False
            self.base_sleep_time = 5  # 重置请求间隔
            self.jd_search_page_loaded = False  # 重置搜索页面状态
            
            # 尝试加载已保存的Cookie
            if os.path.exists(JD_COOKIES_PKL):
                try:
                    with open(JD_COOKIES_PKL, "rb") as f:
                        cookies = pickle.load(f)
                    for cookie in cookies:
                        self.driver.add_cookie(cookie)
                    self.driver.refresh()
                    self.log("✅ 已加载上次登录状态。")
                    # 在主线程中更新UI
                    self.root.update_idletasks()
                    return True
                except Exception as e:
                    self.log("⚠️ 加载 cookie 失败，需要重新登录")
            
            # 需要手动登录
            self.log("🔑 需要用户手动登录京东...")
            messagebox.showinfo("登录提示", "请在打开的浏览器中扫码登录京东，登录完成后点击确定继续。")
            self.driver.get("https://passport.jd.com/new/login.aspx")
            
            # 等待用户登录完成
            user_confirmed = messagebox.askyesno("登录确认", "是否已完成京东登录？\n点击'是'继续，点击'否'重新打开登录页面。")
            while not user_confirmed:
                self.driver.get("https://passport.jd.com/new/login.aspx")
                user_confirmed = messagebox.askyesno("登录确认", "是否已完成京东登录？\n点击'是'继续，点击'否'重新打开登录页面。")
            
            # 登录完成后保存 cookie
            cookies = self.driver.get_cookies()
            with open(JD_COOKIES_PKL, "wb") as f:
                pickle.dump(cookies, f)
            self.log("✅ 登录信息已保存，可下次免登录。")
            # 在主线程中更新UI
            self.root.update_idletasks()
            return True
                
        except Exception as e:
            self.log(f"❌ 京东登录过程出错：{e}")
            return False
    
    def close_browser(self):
        """关闭浏览器"""
        if self.driver:
            try:
                self.driver.quit()
                self.driver = None
                self.log("✅ 浏览器已关闭")
            except Exception as e:
                self.log(f"⚠️ 关闭浏览器时出错：{e}")

    # ---------------- 获取京东价格（使用Selenium） ----------------
    def fetch_price_jd(self, isbn):
        """使用Selenium获取京东价格（仅限自营商品）"""
        if not self.driver:
            self.log("⚠️ 浏览器未初始化")
            return None
            
        try:
            # 随机选择搜索方式：0-直接访问搜索链接，1-在已打开页面中搜索
            import random
            search_method = random.randint(0, 1)
            
            # 每50次请求强制切换一次搜索方式，避免单一模式被识别
            if self.request_count > 0 and self.request_count % 20 == 0:
                search_method = 1 - search_method  # 切换搜索方式
                self.log(f"🔄 切换搜索方式至 {'输入框搜索' if search_method == 1 else '直接链接搜索'}")
            
            if search_method == 0:
                # 方法1：直接访问搜索链接（模拟用户直接在地址栏输入或点击链接）
                self.log(f"🔗 直接访问搜索链接: {isbn}")
                url = f"https://search.jd.com/Search?keyword={isbn}"
                self.driver.get(url)
                
                # 检查是否被跳转到登录页
                current_url = self.driver.current_url
                if self.is_redirected_to_login(current_url):
                    self.log("⚠️ 检测到被跳转到登录页，需要重新登录")
                    if self.handle_cookie_expiration():
                        # 重新登录成功后，重新访问搜索页
                        self.driver.get(url)
                    else:
                        self.log("❌ 重新登录失败")
                        return "登录失败"
            else:
                # 方法2：在已打开页面中搜索（模拟用户在搜索框中输入并点击搜索）
                # 如果是第一次访问或页面未加载，先访问京东搜索页
                if not self.jd_search_page_loaded:
                    self.log("🔍 首次访问京东搜索页...")
                    url = "https://search.jd.com/Search?keyword=978755968178"
                    self.driver.get(url)
                    
                    # 检查是否被跳转到登录页
                    current_url = self.driver.current_url
                    if self.is_redirected_to_login(current_url):
                        self.log("⚠️ 首次访问时检测到被跳转到登录页，需要重新登录")
                        if self.handle_cookie_expiration():
                            # 重新登录成功后，重新访问搜索页
                            url = "https://search.jd.com/Search?keyword=978755968178"
                            self.driver.get(url)
                        else:
                            self.log("❌ 重新登录失败")
                            return "登录失败"
                    
                    # 等待页面加载完成
                    try:
                        WebDriverWait(self.driver, 10).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, "body"))
                        )
                        # 等待搜索框出现
                        WebDriverWait(self.driver, 10).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, "input._search_input_3278r_19"))
                        )
                        self.jd_search_page_loaded = True
                        self.log("✅ 京东搜索页面加载完成")
                    except TimeoutException:
                        self.log("⚠️ 京东搜索页面加载超时")
                        return "超时"
                
                # 检查页面是否仍然有效
                if not self.is_page_content_normal():
                    self.log("⚠️ 检测到页面内容异常，重新加载搜索页面")
                    self.jd_search_page_loaded = False
                    # 重新加载页面
                    url = "https://search.jd.com/Search?keyword=978755968178"
                    self.driver.get(url)
                    
                    # 检查是否被跳转到登录页
                    current_url = self.driver.current_url
                    if self.is_redirected_to_login(current_url):
                        self.log("⚠️ 重新加载时检测到被跳转到登录页，需要重新登录")
                        if self.handle_cookie_expiration():
                            # 重新登录成功后，重新访问搜索页
                            url = "https://search.jd.com/Search?keyword=978755968178"
                            self.driver.get(url)
                        else:
                            self.log("❌ 重新登录失败")
                            return "登录失败"
                    
                    try:
                        WebDriverWait(self.driver, 10).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, "body"))
                        )
                        WebDriverWait(self.driver, 10).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, "input._search_input_3278r_19"))
                        )
                        self.jd_search_page_loaded = True
                        self.log("✅ 京东搜索页面重新加载完成")
                    except TimeoutException:
                        self.log("⚠️ 京东搜索页面重新加载超时")
                        return "超时"
                
                # 检查是否被限制访问（被跳转到首页）
                if self.is_redirected_to_homepage():
                    self.log("⚠️ 检测到被跳转到首页，重新加载搜索页面")
                    self.jd_search_page_loaded = False
                    # 重新加载页面
                    url = "https://search.jd.com/Search?keyword=978755968178"
                    self.driver.get(url)
                    
                    # 检查是否被跳转到登录页
                    current_url = self.driver.current_url
                    if self.is_redirected_to_login(current_url):
                        self.log("⚠️ 重新加载时检测到被跳转到登录页，需要重新登录")
                        if self.handle_cookie_expiration():
                            # 重新登录成功后，重新访问搜索页
                            url = "https://search.jd.com/Search?keyword=978755968178"
                            self.driver.get(url)
                        else:
                            self.log("❌ 重新登录失败")
                            return "登录失败"
                    
                    try:
                        WebDriverWait(self.driver, 10).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, "body"))
                        )
                        WebDriverWait(self.driver, 10).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, "input._search_input_3278r_19"))
                        )
                        self.jd_search_page_loaded = True
                        self.log("✅ 京东搜索页面重新加载完成")
                    except TimeoutException:
                        self.log("⚠️ 京东搜索页面重新加载超时")
                        return "超时"
                
                # 在搜索框中输入ISBN并点击搜索
                try:
                    # 查找搜索输入框
                    search_input = WebDriverWait(self.driver, 10).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "input._search_input_3278r_19"))
                    )
                    
                    # 清空输入框并输入新的ISBN
                    search_input.clear()
                    search_input.send_keys(isbn)
                    
                    # 模拟人工输入后的等待时间（1-3秒）
                    import random
                    wait_time = random.uniform(1, 3)
                    time.sleep(wait_time)
                    
                    # 查找搜索按钮并点击
                    search_btn = WebDriverWait(self.driver, 10).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "button._search_btn_3278r_180"))
                    )
                    search_btn.click()
                    
                    self.log(f"🔍 搜索 ISBN: {isbn}")
                except TimeoutException:
                    self.log("⚠️ 搜索框或搜索按钮未找到")
                    return "超时"
                except Exception as e:
                    self.log(f"⚠️ 搜索操作异常：{e}")
                    return "超时"
            
            # 增加请求计数
            self.request_count += 1
            
            # 等待搜索结果页面加载
            if not self.wait_for_search_results():
                self.log("⚠️ 搜索结果页面加载超时")
                return "超时"
            
            # 检查是否被跳转到登录页（搜索后检查）
            current_url = self.driver.current_url
            if self.is_redirected_to_login(current_url):
                self.log("⚠️ 检测到被跳转到登录页，需要重新登录")
                if self.handle_cookie_expiration():
                    # 重新登录成功后，重新加载搜索页面
                    self.jd_search_page_loaded = False
                    # 重新访问搜索页
                    if search_method == 0:
                        url = f"https://search.jd.com/Search?keyword={isbn}"
                        self.driver.get(url)
                    else:
                        url = "https://search.jd.com/Search?keyword=978755968178"
                        self.driver.get(url)
                        try:
                            WebDriverWait(self.driver, 10).until(
                                EC.presence_of_element_located((By.CSS_SELECTOR, "body"))
                            )
                            WebDriverWait(self.driver, 10).until(
                                EC.presence_of_element_located((By.CSS_SELECTOR, "input._search_input_3278r_19"))
                            )
                            self.jd_search_page_loaded = True
                            self.log("✅ 京东搜索页面重新加载完成")
                            # 重新执行搜索
                            try:
                                # 查找搜索输入框
                                search_input = WebDriverWait(self.driver, 10).until(
                                    EC.presence_of_element_located((By.CSS_SELECTOR, "input._search_input_3278r_19"))
                                )
                                
                                # 清空输入框并输入新的ISBN
                                search_input.clear()
                                search_input.send_keys(isbn)
                                
                                # 模拟人工输入后的等待时间（1-3秒）
                                import random
                                wait_time = random.uniform(1, 3)
                                time.sleep(wait_time)
                                
                                # 查找搜索按钮并点击
                                search_btn = WebDriverWait(self.driver, 10).until(
                                    EC.element_to_be_clickable((By.CSS_SELECTOR, "button._search_btn_3278r_180"))
                                )
                                search_btn.click()
                                
                                self.log(f"🔍 重新搜索 ISBN: {isbn}")
                                # 增加请求计数
                                self.request_count += 1
                                
                                # 等待搜索结果页面加载
                                if not self.wait_for_search_results():
                                    self.log("⚠️ 重新搜索后结果页面加载超时")
                                    return "超时"
                            except TimeoutException:
                                self.log("⚠️ 重新搜索时搜索框或搜索按钮未找到")
                                return "超时"
                            except Exception as e:
                                self.log(f"⚠️ 重新搜索操作异常：{e}")
                                return "超时"
                        except TimeoutException:
                            self.log("⚠️ 重新加载搜索页面超时")
                            return "超时"
                else:
                    return "登录失败"
            
            # 模拟用户操作，防止被识别为自动化脚本
            if self.driver:
                # 随机决定使用基础操作还是高级操作
                import random
                if random.random() < 0.3:  # 30%概率使用高级操作
                    self.simulate_advanced_user_actions()
                else:
                    self.simulate_user_actions()
                
                # 添加模拟人类行为
                if random.random() < 0.4:  # 40%概率模拟人类行为
                    self.simulate_human_behavior()
            # 等待页面基本加载
            try:
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "body"))
                )
                # 增加额外等待时间，确保动态内容完全加载
                time.sleep(5)  # 增加等待时间到5秒

                
                # 重试机制：如果价格提取失败，再等待一段时间后重试
                retry_count = 3
                price = None
                for attempt in range(retry_count):
                    price = self.extract_self_sold_price()
                    # 检查价格是否是不完整的（有小数点但小数点后没有数字）
                    if price and '.' in price and len(price.split('.')[-1]) == 0:
                        # 价格不完整（如"54."），需要重试
                        if attempt < retry_count - 1:
                            self.log(f"⚠️ 价格 {price} 不完整，第{attempt + 1}次重试...")
                            time.sleep(3)  # 等待3秒后重试
                            continue
                        else:
                            # 重试后仍然不完整，返回当前价格
                            self.log(f"⚠️ 价格 {price} 仍然不完整")
                            return price
                    elif price:
                        # 价格完整（包括整数价格如"249"和完整小数价格如"54.5"）
                        if '.' in price and len(price.split('.')[-1]) > 0:
                            self.log(f"✅ 获取到完整价格: {price}")
                        else:
                            self.log(f"✅ 获取到整数价格: {price}")
                        return price
                    elif attempt < retry_count - 1:
                        # 如果没有获取到价格且还有重试机会，等待后重试
                        self.log(f"⚠️ 未获取到价格，第{attempt + 1}次重试...")
                        time.sleep(3)  # 等待3秒后重试
                
                # 如果重试后仍然没有获取到完整价格，返回最后一次的结果
                if price:
                    self.log(f"✅ 获取到价格: {price}")
                return price if price else "超时"
            except TimeoutException:
                self.log("⚠️ 页面加载超时")
                return "超时"
        except Exception as e:
            self.log(f"⚠️ 京东请求异常：{e}")
            return None
    
    def wait_for_search_results(self):
        """等待搜索结果页面加载完成"""
        # 添加类型检查
        if not self.driver:
            return False
            
        try:
            # 等待页面基本元素加载
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "body"))
            )
            
            # 等待搜索结果容器出现
            WebDriverWait(self.driver, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "._wrapper_f6icl_11, .goods-list-v2"))
            )
            
            return True
        except TimeoutException:
            return False
    
    def is_page_content_normal(self):
        """检查页面内容是否正常（未被反爬虫机制拦截）"""
        # 添加类型检查
        if not self.driver:
            return False
            
        try:
            current_url = self.driver.current_url
            
            # 首先检查是否被跳转到登录页
            if self.is_redirected_to_login(current_url):
                return False
            
            # 检查页面标题
            title = self.driver.title
            
            # 如果标题包含"验证码"、"验证"、"拦截"等关键词，则页面异常
            suspicious_keywords = ["验证码", "验证", "拦截", "检测", "安全", "风控", "请稍后"]
            for keyword in suspicious_keywords:
                if keyword in title:
                    return False
            
            # 检查页面内容
            page_source = self.driver.page_source.lower()
            
            # 如果页面包含这些关键词，说明可能被拦截了
            suspicious_contents = [
                "验证码", "验证", "拦截", "检测", "安全验证", 
                "请稍后重试", "请求过于频繁", "访问异常",
                "robot", "spider", "crawler", "blocked"
            ]
            
            for content in suspicious_contents:
                if content in page_source:
                    return False
            
            # 检查是否有搜索框和搜索按钮
            try:
                search_input = self.driver.find_element(By.CSS_SELECTOR, "input._search_input_3278r_19")
                search_btn = self.driver.find_element(By.CSS_SELECTOR, "button._search_btn_3278r_180")
                # 如果能找到搜索框和搜索按钮，说明页面正常
                if search_input and search_btn:
                    return True
            except NoSuchElementException:
                # 如果找不到搜索框或搜索按钮，页面可能异常
                pass
            
            # 检查是否有商品列表容器
            try:
                search_elements = self.driver.find_elements(By.CSS_SELECTOR, "._wrapper_f6icl_11")
                main_search_container = self.driver.find_elements(By.CSS_SELECTOR, ".goods-list-v2")
                
                # 如果页面标题正常且存在搜索结果容器，则页面正常
                if search_elements or main_search_container:
                    return True
                    
                # 如果标题正常但没有搜索结果容器，检查是否是搜索无结果页面
                no_result_element = self.driver.find_elements(By.CSS_SELECTOR, ".search-noresult")
                if no_result_element:
                    # 搜索无结果页也是正常的
                    return True
            except NoSuchElementException:
                pass
            
            # 如果标题正常，且没有明显的拦截内容，则认为页面正常
            return True
            
        except Exception:
            # 出现异常时，默认认为页面正常（避免误判）
            return True
    
    def is_redirected_to_homepage(self):
        """检测是否被跳转到首页（访问受限）"""
        # 添加类型检查
        if not self.driver:
            return False
            
        try:
            current_url = self.driver.current_url
            # 检查是否被跳转到京东首页
            if current_url == "https://www.jd.com/" or current_url == "https://www.jd.com":
                return True
            
            # 检查页面标题是否为京东首页标题
            try:
                title = self.driver.title
                jd_homepage_titles = [
                    "京东(JD.COM)-正品低价、品质保障、配送及时、轻松购物！",
                    "京东-正品低价、品质保障、配送及时、轻松购物！",
                    "京东"
                ]
                if title in jd_homepage_titles:
                    return True
            except:
                pass
            
            # 检查页面内容特征
            try:
                # 检查是否存在首页特有的元素
                self.driver.find_element(By.CSS_SELECTOR, "div#header")  # 首页头部
                self.driver.find_element(By.CSS_SELECTOR, "div#navitems")  # 导航栏
                
                # 检查是否缺少搜索结果页特有的元素
                search_elements = self.driver.find_elements(By.CSS_SELECTOR, "._wrapper_f6icl_11")
                main_search_container = self.driver.find_elements(By.CSS_SELECTOR, ".goods-list-v2")
                
                # 如果既存在首页元素，又缺少搜索结果元素，则可能是被跳转了
                if not search_elements and not main_search_container:
                    return True
            except NoSuchElementException:
                # 如果找不到搜索页特有的元素，可能被跳转了
                pass
            
            return False
        except Exception:
            # 出现异常时，默认不认为是跳转到首页
            return False
    
    def is_redirected_to_login(self, current_url):
        """检测是否被跳转到登录页"""
        # 添加类型检查
        if not self.driver:
            return False
            
        # 转换为小写进行比较
        current_url_lower = current_url.lower()
        
        login_indicators = [
            'passport.jd.com',
            'login.jd.com', 
            '/login',
            'auth.jd.com',
            'signin.jd.com'
        ]
        
        for indicator in login_indicators:
            if indicator in current_url_lower:
                return True
        
        # 检查页面标题是否包含登录相关关键词
        try:
            title = self.driver.title.lower()
            login_titles = ["登录", "sign in", "login", "账户登录"]
            for login_title in login_titles:
                if login_title in title:
                    return True
        except:
            pass
        
        # 检查页面内容是否包含登录相关元素
        try:
            # 检查是否存在登录表单
            login_forms = self.driver.find_elements(By.CSS_SELECTOR, "form[action*='login'], form[id*='login'], .login-form")
            if login_forms:
                return True
                
            # 检查是否存在登录按钮或扫码登录元素
            login_elements = self.driver.find_elements(By.CSS_SELECTOR, 
                "a[href*='login'], button[type='submit'], .login-btn, .qr-login, .login-tab, .login-wrap")
            if login_elements:
                return True
                
            # 检查页面文本是否包含登录关键词
            page_text = self.driver.find_element(By.TAG_NAME, "body").text.lower()
            login_texts = ["账户登录", "用户登录", "登录", "扫码登录", "短信登录", "忘记密码"]
            for login_text in login_texts:
                if login_text in page_text:
                    return True
        except NoSuchElementException:
            pass
        except Exception:
            pass
        
        return False
    
    def simulate_user_actions(self):
        """模拟用户操作，防止被识别为自动化脚本"""
        try:
            from selenium.webdriver.common.action_chains import ActionChains
            import random
            
            # 检查驱动是否存在
            if not self.driver:
                return
            
            # 随机等待一段时间，模拟用户思考
            time.sleep(random.uniform(0.5, 2.0))
            
            # 模拟鼠标移动
            actions = ActionChains(self.driver)
            
            # 获取页面和窗口尺寸
            page_width = self.driver.execute_script("return document.body.scrollWidth")
            page_height = self.driver.execute_script("return document.body.scrollHeight")
            window_width = self.driver.execute_script("return window.innerWidth")
            window_height = self.driver.execute_script("return window.innerHeight")
            
            # 使用更安全的方式确定最大坐标范围
            # 确保不会超出页面边界，同时留出一些边距
            safe_margin = 50
            max_x = min(page_width, window_width) - safe_margin
            max_y = min(page_height, window_height) - safe_margin
            
            # 确保最大坐标不小于安全边距
            max_x = max(safe_margin, max_x)
            max_y = max(safe_margin, max_y)
            
            # 随机选择几个点进行鼠标移动
            move_count = random.randint(3, 6)
            for i in range(move_count):
                # 确保坐标在安全范围内
                x = random.randint(safe_margin, max_x)
                y = random.randint(safe_margin, max_y)
                
                try:
                    # 使用move_to_element_with_offset而不是move_by_offset
                    # 这样更不容易出界
                    body = self.driver.find_element(By.TAG_NAME, "body")
                    actions.move_to_element_with_offset(body, x, y).perform()
                except Exception:
                    # 忽略鼠标移动错误，不影响主要流程
                    pass
                finally:
                    # 每次操作后重新创建动作链
                    actions = ActionChains(self.driver)
                time.sleep(random.uniform(0.1, 0.5))
            
            # 模拟页面滚动
            scroll_count = random.randint(2, 5)
            for _ in range(scroll_count):
                # 随机滚动距离
                scroll_distance = random.randint(100, 500)
                self.driver.execute_script(f"window.scrollBy(0, {scroll_distance});")
                time.sleep(random.uniform(0.2, 1.0))
            
            # 滚动到页面底部再回到顶部
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(random.uniform(0.5, 1.5))
            self.driver.execute_script("window.scrollTo(0, 0);")
            time.sleep(random.uniform(0.5, 1.0))
            
            # 尝试点击页面上的非关键元素（如果存在）
            try:
                # 查找一些常见的非关键元素进行点击
                elements = self.driver.find_elements(By.CSS_SELECTOR, "div, span, p")
                if elements:
                    # 随机选择一个元素进行点击
                    element = random.choice(elements[:min(10, len(elements))])
                    # 确保元素在视口中
                    self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
                    time.sleep(random.uniform(0.2, 0.8))
                    # 点击元素
                    actions = ActionChains(self.driver)
                    actions.move_to_element(element).click().perform()
                    time.sleep(random.uniform(0.3, 0.7))
            except Exception:
                # 如果点击失败，不影响主要流程
                pass
            
            # 模拟用户阅读时间
            time.sleep(random.uniform(1.0, 3.0))
            
            self.log("🖱️ 已模拟用户操作，降低被识别风险")
            
        except Exception as e:
            # 即使模拟操作失败，也不影响主要流程
            # self.log(f"⚠️ 模拟用户操作时出错: {e}")
            pass
    
    def simulate_advanced_user_actions(self):
        """模拟更高级的用户操作，进一步降低被识别风险"""
        try:
            import random
            
            # 检查驱动是否存在
            if not self.driver:
                return
            
            # 随机选择一种高级操作
            action_type = random.randint(1, 4)
            
            if action_type == 1:
                # 模拟鼠标悬停在某些元素上
                try:
                    elements = self.driver.find_elements(By.CSS_SELECTOR, "a, button, ._goodsContainer_f6icl_1 plugin_goodsContainer ._wrapper_8v3rv_3 plugin_goodsCardWrapper _row_6_8v3rv_13")
                    if elements:
                        # 随机选择几个元素进行悬停
                        hover_count = random.randint(2, 5)
                        for _ in range(hover_count):
                            element = random.choice(elements[:min(10, len(elements))])
                            from selenium.webdriver.common.action_chains import ActionChains
                            actions = ActionChains(self.driver)
                            actions.move_to_element(element).perform()
                            time.sleep(random.uniform(0.5, 2.0))
                except Exception:
                    pass
                    
            elif action_type == 2:
                # 模拟键盘操作（如按下和释放Tab键）
                try:
                    from selenium.webdriver.common.keys import Keys
                    body = self.driver.find_element(By.TAG_NAME, "body")
                    body.send_keys(Keys.TAB)
                    time.sleep(random.uniform(0.5, 1.5))
                except Exception:
                    pass
                    
            elif action_type == 3:
                # 模拟更复杂的滚动模式
                try:
                    # 随机滚动到页面中间某个位置
                    page_height = self.driver.execute_script("return document.body.scrollHeight")
                    scroll_position = random.randint(page_height // 4, 3 * page_height // 4)
                    self.driver.execute_script(f"window.scrollTo(0, {scroll_position});")
                    time.sleep(random.uniform(1.0, 3.0))
                    
                    # 再滚动回顶部
                    self.driver.execute_script("window.scrollTo(0, 0);")
                    time.sleep(random.uniform(0.5, 1.5))
                except Exception:
                    pass
                    
            else:
                # 调用基础的用户操作模拟
                self.simulate_user_actions()
                
        except Exception:
            # 即使高级操作失败，也不影响主要流程
            pass
    
    def handle_cookie_expiration(self):
        """处理cookie过期问题"""
        try:
            # 删除过期cookie文件
            if os.path.exists(JD_COOKIES_PKL):
                os.remove(JD_COOKIES_PKL)
            
            # 清理浏览器中cookie
            if self.driver is not None:  # 添加类型检查
                self.driver.delete_all_cookies()
            
            # 重新登录
            return self.login_or_load_jd_cookie()
            
        except Exception as e:
            self.log(f"⚠️ 处理cookie过期失败: {e}")
            return False
    
    def handle_access_restriction(self):
        """处理访问受限情况"""
        self.access_restricted = True
        self.log("🔒 检测到访问受限，增加请求间隔并提示用户")
        
        # 增加基础请求间隔
        self.base_sleep_time = min(self.base_sleep_time + 10, 120)  # 最多增加到120秒
        
        # 显示提示信息
        self.log(f"⚠️ 访问频率可能过高，已自动增加请求间隔至 {self.base_sleep_time} 秒")
        self.log("💡 建议：如持续受限，可尝试以下方法：")
        self.log("   1. 更换网络环境")
        self.log("   2. 稍后再试（建议间隔1小时以上）")
        self.log("   3. 检查京东账户是否正常")
        
        return True
    
    def extract_self_sold_price(self):
        """提取京东自营商品价格 - 根据实际页面结构修复"""
        driver = self.driver
        
        # 添加类型检查
        if driver is None:
            self.log("⚠️ 浏览器未初始化")
            return "错误"
        
        try:
            # 等待商品列表加载 - 使用您提供的实际选择器
            # self.log("⏳ 等待商品列表加载...")
             
            # 使用您提供的实际商品容器选择器
            container_selector = "._wrapper_f6icl_11"
            
            try:
                WebDriverWait(driver, 15).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, container_selector))
                )
                containers = driver.find_elements(By.CSS_SELECTOR, container_selector)
                # self.log(f"✅ 找到 {len(containers)} 个商品容器")
            except TimeoutException:
                # self.log("⚠️ 等待商品容器超时")
                return "超时"
            
            if not containers:
                # self.log("⚠️ 未找到商品容器")
                return "超时"
            
            # 只处理第一个容器（真正的搜索结果），但要遍历其中所有商品
            main_container = containers[0]
            
            # 在主容器中查找所有商品项
            # 尝试多种选择器来查找商品项
            item_selectors = [
                "li[data-sku]",  # 原始选择器
                "li",            # 通用li元素
                "div[data-sku]", # div类型的商品
                ".gl-item",      # 商品项class
                "[class*='item']", # 包含item的class
            ]
            
            items = []
            for selector in item_selectors:
                try:
                    items = main_container.find_elements(By.CSS_SELECTOR, selector)
                    if items:
                        # self.log(f"✅ 使用选择器 '{selector}' 在主容器中找到 {len(items)} 个商品项")
                        break
                    else:
                        # self.log(f"⚠️ 选择器 '{selector}' 未找到商品")
                        pass
                except Exception as e:
                    # self.log(f"⚠️ 选择器 '{selector}' 查找失败: {e}")
                    continue
            
            if not items:
                # self.log("⚠️ 所有选择器都未找到商品项，尝试直接从容器中获取所有子元素")
                try:
                    # 最后的后备方案：获取所有直接子元素
                    items = main_container.find_elements(By.XPATH, "./*")
                    # self.log(f"✅ 后备方案找到 {len(items)} 个子元素")
                except Exception as e:
                    # self.log(f"⚠️ 后备方案也失败: {e}")
                    return "超时"


            # 遍历商品，查找自营商品
            for i, item in enumerate(items):
                try:
                    # 检查是否为自营商品 - 使用您提供的实际结构
                    # self.log(f"🔍 检查商品 #{i+1} 是否为自营...")
                    
                    # 使用您提供的自营标签结构
                    self_support_selector = 'div._imgTag_1qbwk_1 img[alt="自营"]'
                    
                    try:
                        tag = item.find_element(By.CSS_SELECTOR, self_support_selector)
                        if tag:
                            # self.log(f"✅ 找到自营商品 #{i+1}")
                            
                            # 提取价格 - 使用您提供的实际价格结构
                            price = self.extract_price_from_item(item, i+1)
                            if price:
                                return price
                    except NoSuchElementException:
                        # self.log(f"⚠️ 商品 #{i+1} 不是自营")
                        continue
                                
                except Exception as e:
                    # self.log(f"⚠️ 处理商品 #{i+1} 时出错: {e}")
                    continue

            # self.log("⚠️ 未找到自营商品或价格")
            return "无自营"
            
        except Exception as e:
            self.log(f"⚠️ 价格提取异常: {e}")
            return "超时"
    
    def extract_price_from_item(self, item, item_num):
        """从单个商品容器中提取价格 - 修复小数部分提取"""
        try:
            # 等待价格元素加载
            time.sleep(1)
            
            # 直接获取价格标签的HTML内容
            try:
                price_container = item.find_element(By.CSS_SELECTOR, "span._price_uqsva_14")
                if price_container:
                    # 获取容器的HTML内容
                    container_html = price_container.get_attribute('innerHTML')
                    if container_html:
                        import re
                        
                        # 移除货币符号标签
                        clean_html = re.sub(r'<i[^>]*>.*?</i>', '', container_html)
                        
                        # 处理有小数部分的情况：如 "249<span>.</span><span>5</span>"
                        decimal_match = re.search(r'(\d+)<span>\.</span><span[^>]*>(\d+)</span>', clean_html)
                        if decimal_match:
                            integer_part = decimal_match.group(1)
                            decimal_part = decimal_match.group(2)
                            full_price = f"{integer_part}.{decimal_part}"
                            return full_price
                        
                        # 处理普通小数格式：如 "249.5"
                        simple_decimal_match = re.search(r'(\d+\.\d+)', clean_html)
                        if simple_decimal_match:
                            return simple_decimal_match.group(1)
                        
                        # 处理整数情况：如 "249"
                        # 先移除所有HTML标签，然后提取数字
                        text_only = re.sub(r'<[^>]+>', '', clean_html).strip()
                        if text_only:
                            # 查找第一个数字序列
                            number_match = re.search(r'(\d+(?:\.\d+)?)', text_only)
                            if number_match:
                                return number_match.group(1)
                        
                        # 最后的后备方案
                        price_matches = re.findall(r'(\d+(?:\.\d+)?)', clean_html)
                        if price_matches:
                            return price_matches[0]
            except NoSuchElementException:
                pass
            
            # 备用方案：获取文本内容
            try:
                price_container = item.find_element(By.CSS_SELECTOR, "span._price_uqsva_14")
                container_text = price_container.text.strip()
                if container_text:
                    # 清理文本内容
                    import re
                    clean_text = container_text.replace("¥", "").replace("￥", "").strip()
                    # 移除多余的空格和换行符
                    clean_text = re.sub(r'\s+', '', clean_text)
                    if clean_text:
                        # 验证是否为有效价格格式
                        if re.match(r'^\d+(?:\.\d+)?$', clean_text):
                            return clean_text
            except NoSuchElementException:
                pass
            
            # 最后的后备方案
            try:
                all_text = item.text
                if all_text:
                    import re
                    # 查找价格模式
                    price_patterns = [
                        r'(\d+\.\d{1,2})',      # 65.99, 65.9
                        r'(\d{2,})'             # 至少两位数字
                    ]
                    
                    for pattern in price_patterns:
                        matches = re.findall(pattern, all_text)
                        if matches:
                            # 返回第一个匹配的价格
                            return matches[0]
            except Exception:
                pass
                
            return None
            
        except Exception as e:
            # self.log(f"⚠️ 商品 #{item_num} 价格提取异常: {e}")
            return None

    # ---------------- 获取当当价格 ----------------
    def fetch_price_dd(self, isbn):
        # 第1步：请求搜索页
        search_url = f"https://search.dangdang.com/?key={isbn}&act=input&filter=0%7C0%7C0%7C0%7C0%7C1%7C0%7C0%7C0%7C0%7C0%7C0%7C0%7C0%7C0"
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                          "(KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36"
        }

        # print(f"📡 请求当当搜索页：{search_url}")
        resp = requests.get(search_url, headers=headers, timeout=15)

        if resp.status_code != 200:
            # print("❌ 搜索页请求失败")
            return {"price": "", "discount": ""}

        html = resp.text
        # 解析 HTML
        soup = BeautifulSoup(html, "html.parser")

        # 查找指定的商品区域
        search_area = soup.find("div", {"id": "search_nature_rg", "dd_name": "普通商品区域"})
        if not search_area:
            # print("⚠️ 未找到商品区域")
            return {"price": "", "discount": ""}

        # 查找第一个商品项
        first_item = search_area.find("li")
        if not first_item:
            # print("⚠️ 未找到商品项")
            return {"price": "", "discount": ""}

        # 检查是否存在"到货通知"标签，如果存在则表示无货
        # 使用类型忽略注释来解决类型检查问题
        arrival_notice = first_item.find("a", {"class": "search_btn_cart", "name": "pdno"})  # type: ignore
        if arrival_notice and "到货通知" in arrival_notice.get_text():  # type: ignore
            # print("📦 商品暂时无货，需要到货通知")
            return {"price": "", "discount": ""}

        # 提取价格
        price_tag = first_item.find("span", {"class": "search_now_price"})  # type: ignore
        price = price_tag.get_text().strip().replace("¥", "").replace("&yen;", "") if price_tag else ""  # type: ignore

        # 提取商品ID
        # 使用类型检查来解决类型问题
        product_id = None
        if hasattr(first_item, 'get'):  # type: ignore
            product_id = first_item.get("sku") if first_item.has_attr("sku") else None  # type: ignore
        if not product_id:
            # 尝试从链接中提取ID
            product_link = first_item.find("a", {"class": "pic"})  # type: ignore
            if product_link and hasattr(product_link, 'get'):  # type: ignore
                href = product_link.get("href", "")  # type: ignore
                # 确保 href 是字符串类型
                if isinstance(href, str):
                    match = re.search(r'product\.dangdang\.com/(\d+)\.html', href)
                    if match:
                        product_id = match.group(1)

        if not product_id:
            # print("⚠️ 未找到商品ID，无法继续获取优惠")
            return {"price": price, "discount": ""}

        # print(f"✅ 当当价格：¥{price}，商品ID：{product_id}")

        # 第2步：请求优惠接口
        promo_url = (
            f"https://search.dangdang.com/Standard/Search/Extend/hosts/api/get_json.php"
            f"?type=promoIcon"
            f"&keys={product_id}"
            f"&url=0%2F%3Fkey%3D{isbn}%26act%3Dinput%26filter%3D0%7C0%7C0%7C0%7C0%7C1%7C0%7C0%7C0%7C0%7C0%7C0%7C0%7C0%7C0"
            f"&c=false&l=7b2eea5e6454245e9e56ac31ae24f124"
        )

        # 提取 cookies
        cookies = resp.cookies.get_dict()
        cookie_header = resp.headers.get("Set-Cookie", "")
        match_passback = re.search(r"search_passback=([^;]+)", cookie_header)
        search_passback = match_passback.group(1) if match_passback else cookies.get("search_passback", "")

        cookie_str = (
            "ddscreen=2; "
            "__permanent_id=20251008120357543381482942643863985; "
            "dest_area=country_id%3D9000%26province_id%3D111%26city_id%3D0%26district_id%3D0%26town_id%3D0; "
            "__visit_id=20251008205927180350921430742688925; "
            "__out_refer=; "
            "__rpm=s_112100.5402553.8.1759928699406%7Cs_112100.5402553.8.1759928702841; "
            "__trace_id=20251008210502894333062959282935496; "
            f"search_passback={search_passback}"
        )

        promo_headers = {
            "Referer": search_url,
            "Cookie": cookie_str,
            "User-Agent": headers["User-Agent"]
        }

        # print(f"📡 请求优惠接口：{promo_url}")
        try:
            promo_resp = requests.get(promo_url, headers=promo_headers, timeout=10)

            if promo_resp.status_code != 200:
                # print("⚠️ 优惠接口请求失败")
                return {"price": price, "discount": ""}

            data = promo_resp.json()
            promo_list = data.get(product_id, [])
            discounts = [item["label_name"] for item in promo_list if item["label_name"] not in ("自营", "券")]
            discount_text = "，".join(discounts) if discounts else "无"

            # print(f"✅ 优惠信息：{discount_text}")
            return {"price": price, "discount": discount_text}
        except Exception as e:
            # print(f"⚠️ 优惠信息获取异常：{e}")
            return {"price": price, "discount": ""}
    # ---------------- 测试方法 ----------------
    def test_jd_access(self):
        """测试京东访问功能"""
        def run_test():
            try:
                self.log("🧪 开始测试京东访问...")
                
                # 记录开始时间
                start_time = time.time()
                
                # 初始化浏览器
                if not self.init_browser():
                    self.log("❌ 测试失败：浏览器启动失败")
                    return
                
                # 登录京东
                if not self.login_or_load_jd_cookie():
                    self.log("❌ 测试失败：京东登录失败")
                    self.close_browser()
                    return
                
                # 重置搜索页面状态
                self.jd_search_page_loaded = False
                
                # 测试ISBN列表
                test_isbns = [
                    "9787575307130",
                    "9787308262453",
                ]
                
                success_count = 0
                for i, isbn in enumerate(test_isbns, 1):
                    self.log(f"🔍 {i}/{len(test_isbns)} 正在测试 ISBN: {isbn}")
                    price = self.fetch_price_jd(isbn)
                    
                    # 检查是否登录失败
                    if price == "登录失败":
                        self.log("⚠️ 测试过程中检测到登录失败")
                        messagebox.showwarning("登录失败", "测试过程中检测到登录失败，请重新登录后再试。")
                        break
                    
                    # 检查是否访问受限
                    if price == "访问受限":
                        self.log("⚠️ 测试过程中检测到访问受限")
                        messagebox.showwarning("访问受限", f"测试过程中检测到京东访问受限，请稍后再试。\n当前请求间隔为 {self.base_sleep_time} 秒。")
                        break
                    
                    if price and price not in ["无自营", "超时"]:
                        self.log(f"✅ 成功获取价格: ¥{price}")
                        success_count += 1
                    else:
                        self.log(f"⚠️ 未获取到价格: {price or '空'}")
                    
                    # 防止请求过频
                    if i < len(test_isbns):
                        self.log("⏳ 等待5秒...")
                        time.sleep(5)
                
                # 计算总耗时
                end_time = time.time()
                elapsed_time = end_time - start_time
                hours, rem = divmod(elapsed_time, 3600)
                minutes, seconds = divmod(rem, 60)
                
                # 测试结果
                if success_count > 0:
                    self.log(f"✅ 测试完成！成功获取 {success_count}/{len(test_isbns)} 个价格")
                    self.log(f"⏱️ 测试总耗时: {int(hours):02d}时{int(minutes):02d}分{int(seconds):02d}秒")
                    messagebox.showinfo("测试结果", f"测试成功！\n成功获取 {success_count}/{len(test_isbns)} 个价格\n测试总耗时: {int(hours):02d}时{int(minutes):02d}分{int(seconds):02d}秒")
                else:
                    self.log("❌ 测试失败：未能获取任何价格")
                    self.log(f"⏱️ 测试总耗时: {int(hours):02d}时{int(minutes):02d}分{int(seconds):02d}秒")
                    messagebox.showwarning("测试结果", f"测试失败！\n未能获取任何价格，请检查网络和登录状态\n测试总耗时: {int(hours):02d}时{int(minutes):02d}分{int(seconds):02d}秒")
                
            except Exception as e:
                self.log(f"❌ 测试过程中发生错误：{e}")
                messagebox.showerror("测试错误", f"测试过程中发生错误：\n{e}")
            finally:
                # 确保关闭浏览器
                self.close_browser()
        
        # 在新线程中运行测试，防止阻塞UI
        threading.Thread(target=run_test, daemon=True).start()

    # ---------------- 测试当当访问方法 ----------------
    def test_dd_access(self):
        """测试当当访问功能"""
        def run_test():
            # 记录开始时间
            start_time = time.time()
            
            try:
                self.log("🧪 开始测试当当访问...")
                
                # 测试ISBN列表
                test_isbns = [
                    "9787513948128",  # 测试书号1
                    "9787229192914"   # 测试书号2
                ]
                
                success_count = 0
                for i, isbn in enumerate(test_isbns, 1):
                    self.log(f"🔍 {i}/{len(test_isbns)} 正在测试 ISBN: {isbn}")
                    result = self.fetch_price_dd(isbn)
                    
                    # 检查是否成功获取价格
                    if result and isinstance(result, dict) and result.get('price'):
                        price = result['price']
                        self.log(f"✅ 成功获取价格: ¥{price}")
                        success_count += 1
                    elif result and isinstance(result, dict) and result.get('price') == "":
                        self.log("⚠️ 商品暂时无货或未找到")
                    else:
                        self.log("⚠️ 未获取到价格信息")
                    
                    # 防止请求过频
                    if i < len(test_isbns):
                        self.log("⏳ 等待3秒...")
                        time.sleep(3)
                
                # 计算总耗时
                end_time = time.time()
                elapsed_time = end_time - start_time
                hours, rem = divmod(elapsed_time, 3600)
                minutes, seconds = divmod(rem, 60)
                
                # 测试结果
                if success_count > 0:
                    self.log(f"✅ 测试完成！成功获取 {success_count}/{len(test_isbns)} 个价格")
                    self.log(f"⏱️ 测试总耗时: {int(hours):02d}时{int(minutes):02d}分{int(seconds):02d}秒")
                    messagebox.showinfo("测试结果", f"测试成功！\n成功获取 {success_count}/{len(test_isbns)} 个价格\n测试总耗时: {int(hours):02d}时{int(minutes):02d}分{int(seconds):02d}秒")
                else:
                    self.log("⚠️ 测试完成：未能获取有效价格")
                    self.log(f"⏱️ 测试总耗时: {int(hours):02d}时{int(minutes):02d}分{int(seconds):02d}秒")
                    messagebox.showwarning("测试结果", f"测试完成！\n未能获取有效价格，请检查网络连接\n测试总耗时: {int(hours):02d}时{int(minutes):02d}分{int(seconds):02d}秒")
                
            except Exception as e:
                self.log(f"❌ 测试过程中发生错误：{e}")
                messagebox.showerror("测试错误", f"测试过程中发生错误：\n{e}")
        
        # 在新线程中运行测试，防止阻塞UI
        threading.Thread(target=run_test, daemon=True).start()
    
    def test_login_detection(self):
        """测试登录检测功能"""
        def run_test():
            try:
                self.log("🧪 开始测试登录检测功能...")
                
                # 初始化浏览器
                if not self.init_browser():
                    self.log("❌ 测试失败：浏览器启动失败")
                    return
                
                # 检查驱动是否初始化
                if not self.driver:
                    self.log("❌ 浏览器驱动未初始化")
                    return
                
                # 访问京东登录页面进行测试
                self.log("🔍 访问京东登录页面进行测试...")
                self.driver.get("https://passport.jd.com/new/login.aspx")
                
                # 获取当前URL和标题
                current_url = self.driver.current_url
                title = self.driver.title
                
                # 检查是否能正确检测到登录页
                is_login_page = self.is_redirected_to_login(current_url)
                if is_login_page:
                    self.log("✅ 成功检测到登录页面")
                    messagebox.showinfo("测试结果", "成功检测到登录页面！\n登录检测功能正常。")
                else:
                    self.log("❌ 未能检测到登录页面")
                    messagebox.showwarning("测试结果", "未能检测到登录页面！\n登录检测功能可能存在问题。")
                
            except Exception as e:
                self.log(f"❌ 测试过程中发生错误：{e}")
                messagebox.showerror("测试错误", f"测试过程中发生错误：\n{e}")
            finally:
                # 确保关闭浏览器
                self.close_browser()
        
        # 在新线程中运行测试，防止阻塞UI
        threading.Thread(target=run_test, daemon=True).start()
    
    def reset_access_interval(self):
        """重置访问间隔"""
        self.base_sleep_time = 5
        self.access_restricted = False
        self.request_count = 0
        self.log("🔄 已重置访问间隔为20秒")
        messagebox.showinfo("重置成功", "已将访问间隔重置为20秒")
    
    def calculate_sleep_time(self):
        """根据请求次数动态计算睡眠时间"""
        import random
        
        # 基础间隔时间
        base_time = self.base_sleep_time
        
        # 根据请求次数增加间隔时间
        # 每20次请求增加5秒间隔，最多增加到60秒
        additional_time = min((self.request_count // 20) * 5, 40)
        
        # 计算总间隔时间
        total_time = base_time + additional_time
        
        # 添加随机波动（±5秒）
        random_variation = random.uniform(-5, 5)
        final_time = max(total_time + random_variation, 5)  # 最少5秒
        
        # 添加额外的随机延迟，模拟真实用户行为
        # 随机增加0-10秒的延迟
        extra_random_delay = random.uniform(0, 10)
        final_time += extra_random_delay
        
        return final_time
    
    def simulate_human_behavior(self):
        """模拟更真实的人类行为，降低被识别为爬虫的概率"""
        # 检查驱动是否存在
        if not self.driver:
            return
            
        try:
            import random
            from selenium.webdriver.common.action_chains import ActionChains
            
            # 随机选择行为类型
            behavior_type = random.randint(1, 5)
            
            if behavior_type == 1:
                # 模拟鼠标轨迹移动
                self._simulate_mouse_movement()
            elif behavior_type == 2:
                # 模拟页面滚动
                self._simulate_page_scrolling()
            elif behavior_type == 3:
                # 模拟键盘输入
                self._simulate_keyboard_input()
            elif behavior_type == 4:
                # 模拟元素悬停
                self._simulate_element_hover()
            else:
                # 综合行为
                self._simulate_combined_behavior()
                
            self.log("👤 已模拟人类行为，降低被识别风险")
            
        except Exception as e:
            # 即使模拟行为失败，也不影响主要流程
            pass
    
    def _simulate_mouse_movement(self):
        """模拟鼠标移动轨迹"""
        # 检查驱动是否存在
        if not self.driver:
            return
            
        try:
            import random
            from selenium.webdriver.common.action_chains import ActionChains
            
            actions = ActionChains(self.driver)
            
            # 获取页面尺寸
            page_width = self.driver.execute_script("return document.body.scrollWidth")
            page_height = self.driver.execute_script("return document.body.scrollHeight")
            
            # 生成鼠标移动轨迹点
            points = []
            num_points = random.randint(5, 15)
            
            for i in range(num_points):
                x = random.randint(0, page_width - 1)
                y = random.randint(0, page_height - 1)
                points.append((x, y))
            
            # 执行鼠标移动
            body = self.driver.find_element(By.TAG_NAME, "body")
            for x, y in points:
                actions.move_to_element_with_offset(body, x, y).perform()
                time.sleep(random.uniform(0.05, 0.2))
                
        except Exception:
            pass
    
    def _simulate_page_scrolling(self):
        """模拟页面滚动行为"""
        # 检查驱动是否存在
        if not self.driver:
            return
            
        try:
            import random
            
            # 随机选择滚动模式
            scroll_mode = random.randint(1, 3)
            
            if scroll_mode == 1:
                # 缓慢滚动到页面底部
                page_height = self.driver.execute_script("return document.body.scrollHeight")
                current_pos = 0
                while current_pos < page_height:
                    scroll_step = random.randint(50, 200)
                    current_pos += scroll_step
                    self.driver.execute_script(f"window.scrollTo(0, {min(current_pos, page_height)});")
                    time.sleep(random.uniform(0.1, 0.5))
            elif scroll_mode == 2:
                # 随机滚动
                for _ in range(random.randint(3, 8)):
                    scroll_distance = random.randint(-300, 300)
                    self.driver.execute_script(f"window.scrollBy(0, {scroll_distance});")
                    time.sleep(random.uniform(0.2, 1.0))
            else:
                # 滚动到特定位置再返回
                page_height = self.driver.execute_script("return document.body.scrollHeight")
                target_pos = random.randint(page_height // 4, 3 * page_height // 4)
                self.driver.execute_script(f"window.scrollTo(0, {target_pos});")
                time.sleep(random.uniform(1.0, 3.0))
                self.driver.execute_script("window.scrollTo(0, 0);")
                
        except Exception:
            pass
    
    def _simulate_keyboard_input(self):
        """模拟键盘输入行为"""
        # 检查驱动是否存在
        if not self.driver:
            return
            
        try:
            import random
            from selenium.webdriver.common.keys import Keys
            
            # 随机选择是否进行键盘输入
            if random.random() < 0.3:  # 30%概率进行键盘输入
                body = self.driver.find_element(By.TAG_NAME, "body")
                
                # 模拟按下Tab键
                body.send_keys(Keys.TAB)
                time.sleep(random.uniform(0.5, 1.5))
                
                # 模拟按下箭头键
                arrow_keys = [Keys.ARROW_DOWN, Keys.ARROW_UP, Keys.ARROW_LEFT, Keys.ARROW_RIGHT]
                for _ in range(random.randint(1, 3)):
                    key = random.choice(arrow_keys)
                    body.send_keys(key)
                    time.sleep(random.uniform(0.2, 0.8))
                    
        except Exception:
            pass
    
    def _simulate_element_hover(self):
        """模拟元素悬停行为"""
        # 检查驱动是否存在
        if not self.driver:
            return
            
        try:
            import random
            from selenium.webdriver.common.action_chains import ActionChains
            
            # 查找页面上的元素
            elements = self.driver.find_elements(By.CSS_SELECTOR, "a, button, img, div")
            if elements:
                actions = ActionChains(self.driver)
                
                # 随机选择几个元素进行悬停
                hover_count = random.randint(2, 5)
                for _ in range(hover_count):
                    element = random.choice(elements[:min(10, len(elements))])
                    try:
                        actions.move_to_element(element).perform()
                        time.sleep(random.uniform(0.5, 2.0))
                    except Exception:
                        pass
                        
        except Exception:
            pass
    
    def _simulate_combined_behavior(self):
        """模拟综合行为"""
        # 检查驱动是否存在
        if not self.driver:
            return
            
        try:
            import random
            
            # 组合多种行为
            behaviors = [
                self._simulate_mouse_movement,
                self._simulate_page_scrolling,
                self._simulate_keyboard_input,
                self._simulate_element_hover
            ]
            
            # 随机执行1-3种行为
            selected_behaviors = random.sample(behaviors, random.randint(1, 3))
            for behavior in selected_behaviors:
                behavior()
                time.sleep(random.uniform(0.5, 1.5))
                
        except Exception:
            pass

if __name__ == "__main__":
    root = tk.Tk()
    app = JDPriceFetcherApp(root)
    root.mainloop()
